import random
import io
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.pagesizes import A4
import ast
from barcode.writer import ImageWriter
import barcode
import base64
from calendar import month_abbr, month_name
from datetime import timedelta
import xlwt
from django.http import HttpResponse
from openpyxl import load_workbook
import os
from PIL import Image
from io import BytesIO
from django.core.files import File
import operator
import functools
import collections
from collections import OrderedDict
from datetime import datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.generics import CreateAPIView
from rest_framework import status
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate, login
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth import get_user_model

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from django.core.mail import EmailMessage
from . import models, serializers
from django.core.exceptions import ValidationError
import json


def CHECK_IN_PLAN_AND_RESPONSE(user, data, **args):
    if user.is_plan:
        return Response('End Plan or No Purchase Plan')
    else:
        return Response(data=data, **args)

    print('User is in Plan')


class LoginView(APIView):

    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):

        print(request.data)

        username_or_email = request.data['username']
        password = request.data['password']
        device_unique = request.data.get("unique_id", None)
        device_name = request.data.get("device_name", None)
        acc_type = request.data.get("acc_type", "Cashier")

        # user devices
        device = request.data.get('device', None)

        user = None
        if '@' in username_or_email:
            b = models.User.objects.get(email=username_or_email)
            user = authenticate(username=b.username, password=password)

        else:
            user = authenticate(username=username_or_email, password=password)

        if not user:
            return Response({'error': 'Invalid Credentials'},
                            status=status.HTTP_401_UNAUTHORIZED)

        # user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)

        device_count = models.Device.objects.filter(user=user).count()

        if device_count > user.device_limit:
            raise ValidationError('Device Limit Exceeded')
        else:
            models.Device.objects.create(
                user=user, unique_id=device_unique, device_name=device_name, acc_type=acc_type)

        # add custom data to response
        response_data = {
            'token': token.key,
            'user_id': user.id,
            'username': user.username,
            'email': user.email,
            # add additional data fields here as needed
        }

        return Response(response_data)


class CreateUserApiView(CreateAPIView):

    permission_classes = [AllowAny]
    serializer_class = serializers.CreateUserSerializer

    def post(self, request):
        print('posteed')
        print(request.data)

        device_unique = request.data.get("uid", None)
        device_name = request.data.get("device_name", None)
        acc_type = request.data.get("acc_type", "Cashier")

        serializers = self.get_serializer(data=request.data)
        serializers.is_valid(raise_exception=True)
        self.perform_create(serializers)
        headers = self.get_success_headers(serializers.data)
        user = models.User.objects.get(username=request.data.get('username'))
        models.Device.objects.create(
            user=user, unique_id=device_unique, device_name=device_name, acc_type=acc_type)
        # Create a token than will be used for future auth
        token = Token.objects.create(user=serializers.instance)
        token_data = {'token': token.key}

        return Response(
            {**serializers.data, **token_data},
            status=status.HTTP_201_CREATED,
            headers=headers)


class password_recovery(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        username = request.data.get('username', None)
        user = get_user_model().objects.get(username=username)
        otp = random.randint(100000, 999999)
        models.OTP.objects.create(user=user, otp=otp)

        subject = "Perfect Solution Password Recovery"

        html_content = '''<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Reset Password</title>
  </head>
  <body>
    <p>Hi ''' + username + '''</p>
    <h1>Perfect Solutions</h1>
    <p>
      We received a request to the reset password associated with your Perfect
      Solutions' POS account.
    </p>
    <p>Here is your OTP code:<br /></p>
    <h1 style="width:100%; text-align:center">'''+str(otp)+'''</h1>
    <p>Thank you.</p>
  </body>
</html>
        '''
        SENDEMAIL = EmailMessage(subject, html_content, 'perfectsolutionpos@gmail.com', [
                                 user.email], headers={'Message-ID': user.id})
        SENDEMAIL.content_subtype = "html"
        SENDEMAIL.send()
        # TODO: Send the password reset link to the user's email address
        # email = ...
        # send_email(email, uid, token)

        return Response({"email": user.email}, status=status.HTTP_200_OK)


class CheckOTPandChangePassword(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        username = request.GET.get('username', None)
        otp = request.GET.get('otp', None)
        print(otp)
        user = get_user_model().objects.get(username=username)
        otp = models.OTP.objects.get(user=user, otp=otp)
        if otp is not None:
            return Response(status=status.HTTP_200_OK)
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        username = request.data.get('username', None)
        otp = request.data.get('otp', None)
        password = request.data.get('password', None)
        oldpassword = request.data.get('oldpassword', None)

        if oldpassword != None:
            user = get_user_model().objects.get(username=request.user)

            if user.check_password(oldpassword):
                user.set_password(password)
                user.save()
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST)

        else:
            user = get_user_model().objects.get(username=username)
            m_otp = models.OTP.objects.get(user=user, otp=otp)

            user.set_password(password)
            user.save()
            m_otp.delete()

            return Response(status=status.HTTP_200_OK)


class ProfileUpdate(APIView):

    def put(self, request):

        print(request.data)
        user = get_user_model().objects.get(username=request.user)
        name = request.data.get('name', None)
        email = request.data.get('email', None)
        phone = request.data.get('phone', None)
        address = request.data.get('address', None)

        if name is not None:
            user.name = name
        if email is not None:
            user.email = email
        if phone is not None:
            user.phoneno = phone

        if address is not None:
            user.address = address

        user.save()
        s = serializers.ProfileSerializer(user)

        return Response(status=status.HTTP_201_CREATED, data=s.data)


class Category(APIView):
    # permission_classes = [AllowAny]

    def get(self, request):
        user = get_user_model().objects.get(username=request.user)
        data = models.Category.objects.filter(user=user)
        s = serializers.CategorySerializer(data, many=True)

        return Response(s.data)

    def post(self, request):
        user = get_user_model().objects.get(username=request.user)
        models.Category.objects.create(title=request.data['title'], user=user)
        return Response(status=status.HTTP_201_CREATED)


def compress_image(image):
    im = Image.open(image)
    size = File(image).size
    if size > 0.3*1024*1024:
        print('Compressing in Progress')
        if im.mode != 'RGB':
            im = im.convert('RGB')
        im_io = BytesIO()
        im.save(im_io, 'jpeg', quality=10, optimize=True)
        print('Compressing Completed')
        new_image = File(im_io, name=image.name)
        return new_image
    return image


class Product(APIView):
    # permission_classes = [AllowAny]

    def get(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)

        data = models.Product.objects.filter(user=user)
        s = serializers.ProductSerializer(data, many=True)
        return Response(s.data)

    def post(self, request):
        name = request.data.get('name', None)
        price = request.data.get('price', None)
        cost = request.data.get('cost', None)
        qty = request.data.get('qty', None)
        barcode = request.data.get('barcode', None)
        description = request.data['description']
        category = models.Category.objects.get(id=request.data['category'])
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        pic = request.data['pic']
        supplier_name = request.data.get('supplier_name', None)


        md = models.Product.objects.create(
            name=name, user=user, pic=pic, price=price,
            cost=cost, qty=qty, description=description,
            barcode=barcode,
            category=category)

        if not supplier_name == None:
            try:
                supp = models.Supplier.objects.get(name=supplier_name, user=user)
                supp.products.add(md)
                supp.save()
            except ObjectDoesNotExist:
                supp = models.Supplier.objects.create(name=supplier_name, user=user)
                supp.products.add(md)
                supp.save()

        if not pic == 'null':
            md.pic = compress_image(pic)
            md.save()

        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):
        id = request.data['id']
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        name = request.data['name']
        price = request.data['price']
        qty = request.data['qty']
        pic = request.data['pic']
        cost = request.data.get('cost')

        description = request.data['description']
        category = models.Category.objects.get(
            id=request.data['category'], user=user)

        PRODUCTS = models.Product.objects.get(user=user, id=id)
        PRODUCTS.name = name
        PRODUCTS.price = price
        PRODUCTS.qty = qty
        PRODUCTS.cost = cost
        # PRODUCTS.date = date
        PRODUCTS.description = description
        PRODUCTS.category = category
        PRODUCTS.barcode = request.data.get('barcode', PRODUCTS.barcode)

        if not pic == 'null':
            PRODUCTS.pic = compress_image(pic)

        PRODUCTS.save()

        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        id = request.data['id']
        user = get_user_model().objects.get(username=request.user, is_plan=True)

        PRODUCTS = models.Product.objects.get(user=user, id=id)
        PRODUCTS.delete()
        return Response(status=status.HTTP_201_CREATED)


class ProductPriceChangeWithPercentage(APIView):
    def put(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        minus_percentage = request.data.get('minus_perctange', None)
        plus_percentage = request.data.get('plus_perctange', None)

        # print(request.data,minus_percentage,plus_percentage)

        # # If minus_perctange is not None then minus_percentage will be applied to all products' price to decrease the price eg: if 5% is given then 5% will be decreased from all products' price
        # print(float(minus_percentage) * 100,"Minus Perctangle")
        if minus_percentage is not None:
            products = models.Product.objects.filter(user=user)
            for product in products:
                print(product.price, "Original ")
                minus_percentage = float(minus_percentage)
                new_price = round(
                    float(product.price) - (float(product.price) * minus_percentage / 100), 2)
                product.price = new_price
                print(product.price, "Minus ")
                product.save()

        if plus_percentage is not None:
            products = models.Product.objects.filter(user=user)
            for product in products:
                new_price = round(
                    float(product.price) + (float(product.price) * float(plus_percentage) / 100), 2)
                product.price = new_price
                product.save()

        return Response(status=status.HTTP_201_CREATED)


class ProductPrice(APIView):
    def get(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        products = models.Product.objects.filter(user=user)

        # Sum all products cost and multiply with qty to calucate purchase
        total_purchase = 0
        for product in products:
            total_purchase += float(product.cost) * float(product.qty)

        return Response(data=total_purchase, status=status.HTTP_201_CREATED)


def yearGenerator(self, data, strftime='%b'):
    # Create a list of all months in the year

    if strftime == '%b':
        months = [month_abbr[i] for i in range(1, 13)]
    else:
        months = [month_name[i] for i in range(1, 13)]

    # Initialize the result dictionary with all months set to 0
    result = {month: 0 for month in months}

    for month in months:
        print(month)

    for item in data:
        d = datetime.strptime(str(item.date)[0:19], "%Y-%m-%d %H:%M:%S")
        month_name_str = d.strftime(strftime)
        print(month_name_str)
        add_price = float(item.totalProfit)
        result[month_name_str] += add_price

    return result


def monthGenerator(self, data):
    result = {}
    # monthString = ['0','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    # 2022-09-06 18:50:44.169216+00:00

    for item in data:
        d = datetime.strptime(str(item.date)[0:19], "%Y-%m-%d %H:%M:%S")
        name = d.strftime('%x')
        result[name] = result.get(name, 0) + int(float(item.grandtotal))

    return result


def todayGenerator(self, data):
    result = {}
    # monthString = ['0','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    # 2022-09-06 18:50:44.169216+00:00

    for item in data:

        d = datetime.strptime(str(item.date)[0:19], "%Y-%m-%d %H:%M:%S")

        name = d.strftime('%I:%M %p')
        result[name] = result.get(name, 0) + int(float(item.grandtotal))

    return result


def ChartGenerator(self, data, time):
    result = {}
    for item in data:
        data_d = item['date']
        print(data_d, 'Data D')
        d = datetime.strptime(str(data_d), '%Y-%m-%dT%H:%M:%S.%f%z')

        print(d)
        name = d.strftime('%x')
        if time == 'today':
            name = d.strftime('%I:%M %p')
        elif time == 'month':
            name = d.strftime('%x')
        elif time == 'year':
            name = d.strftime('%B')
        else:
            name = d.strftime('%x')
        # print(name)
        result[name] = result.get(name, 0) + int(float(item['grandtotal']))

    return result


def taxCalculatorWithPerctange(self, price, tax_percentage):
    tax_percentage = float(tax_percentage)
    tax = round((float(price) * tax_percentage / 100), 2)
    return tax


def discountCalculatorWithPerctange(self, price, discount_percentage):
    discount_percentage = float(discount_percentage)
    discount = round((float(price) * discount_percentage / 100), 2)
    return discount


class Sales(APIView):
    # permission_classes = [AllowAny]P

    def get(self, request):
        type = request.GET.get('type')
        time = request.GET.get('time')
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        d = datetime.now()
        print(d)
        chartdata = {}

        if time == 'today':
            data = models.Sales.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))

        elif time == 'month':
            data = models.Sales.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))

        elif time == 'year':
            data = models.Sales.objects.filter(
                user=user, date__year=str(d.year))

        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            data = models.Sales.objects.filter(user=user, date__range=(sd, ed))

        else:
            data = models.Sales.objects.filter(user=user)

        if type == 'DT':
            s = serializers.DTSalesSerializer(data, many=True)
        else:
            s = serializers.SalesSerializer(data, many=True)

        chartdata = ChartGenerator(self, s.data, time)

        # chartdata = todayGenerator(self, s.data)
        # print(s.data,'\n')
        # print(data[0].date)
        # print(chartdata.keys())

        CombineData = {
            'DATA': s.data,
            'CHART_LABEL': chartdata.keys(),
            'CHART_DATA': chartdata.values(),
            'CHART': chartdata,
        }

        return Response(CombineData)

    def post(self, request):
        customerName = request.data['customerName']
        products = request.data['products']
        totalAmount = request.data['totalAmount']
        tax = request.data['tax']
        discount = request.data['discount']
        grandtotal = request.data['grandtotal']
        description = request.data['description']
        deliveryCharges = request.data.get('deliveryCharges', None)
        isSaveCustomer = request.data.get('isSaveCustomer', False)
        payment_amount =  request.data.get('payment_amount',None)

        # receiptnumber from sales table and plus one

        user = get_user_model().objects.get(username=request.user, is_plan=True)
        last = models.Sales.objects.filter(user=user).last()

        if last is None:
            rn = '1'.zfill(5)
        else:
            rn = str(int(last.voucherNumber) + 1).zfill(5)

        S = models.Sales.objects.create(user=user, customerName=customerName, voucherNumber=rn,
                                        totalAmount=totalAmount, tax=tax, discount=discount, grandtotal=grandtotal,customer_payment=payment_amount,
                                        deliveryCharges=deliveryCharges, description=description)
        
        S.save()

        print(products)
        p = json.loads(products)
        print(p)

        totalProfit = 0

        totalAmount = 0  # Each Product

        for b in p:
            print(b)
            # if product is not have in the databse add new product in the database with named catageory to Extra Products

            try:

                product = models.Product.objects.get(id=b['name'], user=user)
                product.qty = int(product.qty) - int(b['qty'])
                product.save()

                profit = (float(b['price']) -
                          float(product.cost)) * float(b['qty'])
                ta = (float(product.price)) * float(b['qty'])

                print("Profit try : ", profit)

                models.SoldProduct.objects.create(
                    name=b['pdname'], price=b['price'], profit=profit, qty=b['qty'],
                    productid=product.id, sales=S, user=user)

            except ObjectDoesNotExist:
                try:
                    category = models.Category.objects.get(
                        title='Extra Products', user=user)
                    product = models.Product.objects.create(
                        name=b['pdname'], user=user, price=b['price'], cost=int(
                            b['cost']),
                        qty=5, description='Extra Products', category=category)

                    profit = (float(b['price']) -
                              float(product.cost)) * float(b['qty'])
                    ta = (float(product.price)) * float(b['qty'])

                    a = models.SoldProduct.objects.create(
                        name=b['pdname'], price=b['price'], qty=b['qty'],
                        productid=product.id, profit=profit, sales=S, user=user)

                except ObjectDoesNotExist:
                    category = models.Category.objects.create(
                        title='Extra Products', user=user)
                    product = models.Product.objects.create(
                        name=b['pdname'], user=user, price=b['price'], cost=int(
                            b['cost']),
                        qty=5, description='Extra Products', category=category)

                    profit = (float(b['price']) -
                              float(product.cost)) * float(b['qty'])

                    ta = (float(product.price)) * float(b['qty'])

                    a = models.SoldProduct.objects.create(
                        name=b['pdname'], price=b['price'], qty=b['qty'],
                        productid=product.id, profit=profit, sales=S, user=user)

            totalAmount += ta
            totalProfit += profit

        # S.totalProfit = totalProfit - taxCalculatorWithPerctange(self, totalProfit, S.tax)
        S.totalProfit = totalProfit - \
            discountCalculatorWithPerctange(self, totalAmount, S.discount)
        

        S.save()

        saleseri = serializers.SalesSerializer(S)

        if isSaveCustomer == 'true':
            try:
                Customer = models.CustomerName.objects.get(
                    name=customerName, user=user)
                Customer.sales.add(S)
                Customer.save()
            except ObjectDoesNotExist:
                Customer = models.CustomerName.objects.create(
                    name=customerName, user=user)
                Customer.sales.add(S)
                Customer.save()

        return Response(status=status.HTTP_201_CREATED, data=saleseri.data)

    # Delete will be sale return to products and delete from sales

    def delete(self, request):
        id = request.GET.get('id')
        user = get_user_model().objects.get(username=request.user, is_plan=True)

        S = models.Sales.objects.get(user=user, receiptNumber=id)

        Product = models.SoldProduct.objects.filter(sales=S)
        for p in Product:
            product = models.Product.objects.get(id=p.productid, user=user)
            product.qty = int(product.qty) + int(p.qty)
            product.save()
            p.delete()
        S.delete()

        return Response(status=status.HTTP_201_CREATED)

    # if user edit the sales voucher like price and qty then it will be updated in the database
    # if user increase product qty then it will be decrease from the Product models qty and if user decrease product qty then it will be increase from the Product models qty
    # then it will be updated in the SoldProduct Database

    def put(self, request):
        print(request.data)
        id = request.data['id']
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        customerName = request.data['customerName']
        products = request.data['products']

        S = models.Sales.objects.get(user=user, receiptNumber=id)
        S.customerName = customerName
     #   S.totalAmount = totalAmount
        # S.tax = tax
        # S.discount = discount
        # S.grandtotal = grandtotal
        # S.description = description
        # S.deliveryCharges = deliveryCharges
        S.save()

        editedProducts = products

        currentSoldProduct = models.SoldProduct.objects.filter(
            sales=S, user=user)

        pf = 0  # total profit
        ta = 0  # total amount

        for currentProduct in currentSoldProduct:
            found = False
            for editedProduct in editedProducts:
                if currentProduct.id == editedProduct['id']:
                    found = True
                    print("CURRENT PRODUCTID", currentProduct.id)
                    product = models.Product.objects.get(
                        id=currentProduct.productid, user=user)
                    if int(currentProduct.qty) > int(editedProduct['qty']):
                        product.qty = int(
                            product.qty) + (int(currentProduct.qty) - int(editedProduct['qty']))
                        product.save()
                    elif int(currentProduct.qty) < int(editedProduct['qty']):
                        product.qty = int(
                            product.qty) - (int(editedProduct['qty']) - int(currentProduct.qty))
                        product.save()

                    currentProduct.qty = editedProduct['qty']
                    currentProduct.price = editedProduct['price']
                    pf += (float(currentProduct.price) -
                           float(product.cost)) * float(editedProduct['qty'])
                    ta += (float(currentProduct.price)) * \
                        float(editedProduct['qty'])
                    currentProduct.profit = (
                        float(currentProduct.price) - float(product.cost)) * float(editedProduct['qty'])
                    currentProduct.save()

                    break
            if not found:
                product = models.Product.objects.get(
                    id=currentProduct.productid, user=user)
                product.qty += int(currentProduct.qty)
                product.save()
                currentProduct.delete()

        S.totalAmount = ta
        S.grandtotal = ta + taxCalculatorWithPerctange(self, ta, S.tax) - discountCalculatorWithPerctange(
            self, ta, S.discount) + float(S.deliveryCharges)
        S.totalProfit = pf - \
            discountCalculatorWithPerctange(self, ta, S.discount)
        S.save()

        return Response(status=status.HTTP_201_CREATED)


class CustomerView(APIView):

    def get(self, request):
        user = get_user_model().objects.get(username=request.user)
        cu = models.CustomerName.objects.filter(user=user)

        ser = serializers.CustomerSerializer(cu, many=True)
        return Response(ser.data)

    def post(self, request):
        customer_name = request.data.get('customerName', None)
        description = request.data.get('description', None)
        user = get_user_model().objects.get(username=request.user)
        cu = models.CustomerName.objects.create(name=customer_name, description=description,user=user)
        return Response(status=status.HTTP_201_CREATED)

    def put(self,request):
        sale_id =  request.data.get('sale_id',None)
        customer_payment =  request.data.get('customer_payment',None)

        user = get_user_model().objects.get(username=request.user, is_plan=True)


        #add sales to customer
        customer_id = request.data.get("customer_id", None)
        sales_receipt =  request.data.get('sales', None)

        # sales data have only saleReceiptNumber [1, 4, 6, 0]

        if not sales_receipt == None:
            C = models.CustomerName.objects.get(id=customer_id)

            for sale in sales_receipt:
                S = models.Sales.objects.get(user=user, receiptNumber=sale)
                S.customer_payment = 0;
                S.save()
                C.sales.add(S)
                C.save()
            
                # Customer Set Payment
        if not sale_id == None:
            if not customer_payment == 0:
                S = models.Sales.objects.get(user=user, receiptNumber=sale_id)
                S.customer_payment = int(S.customer_payment) + int(customer_payment)
                S.save()

        return Response(status=status.HTTP_200_OK)

    def delete(self,request):

        customer_id = request.GET.get("customerid", None)
        sales_receipt =  request.GET.get('sales', None)

        user = get_user_model().objects.get(username=request.user, is_plan=True)

        print(sales_receipt, customer_id)

        if not sales_receipt == None:
            C = models.CustomerName.objects.get(id=customer_id)
            S = models.Sales.objects.get(user=user,receiptNumber=sales_receipt)
            S.save()
            C.sales.remove(S)
            C.save()
        else:
            C = models.CustomerName.objects.get(id=customer_id)
            C.delete()


        return Response(status=status.HTTP_200_OK)


class SupplierView(APIView):

    def get(self, request):
        user = get_user_model().objects.get(username=request.user)
        sup = models.Supplier.objects.filter(user=user)
        ser = serializers.SupplierSerializer(sup, many=True)
        return Response(ser.data)

    def post(self, request):
        supplier_name = request.data.get('supplierName', None)
        description = request.data.get('description', None)
        user = get_user_model().objects.get(username=request.user)
        cu = models.Supplier.objects.create(name=supplier_name, description=description,user=user)
        return Response(status=status.HTTP_201_CREATED)

    def put(self,request):
        product_id =  request.data.get('product_id',None)
        supplier_payment =  request.data.get('supplier_payment',None)

        user = get_user_model().objects.get(username=request.user, is_plan=True)


        #add sales to customer
        supplier_id = request.data.get("supplier_id", None)
        productss =  request.data.get('products', None)

        

        # productss data have only saleReceiptNumber [1, 4, 6, 0]

        if not productss == None:
            C = models.Supplier.objects.get(id=supplier_id, user=user)

            for pd_id in productss:
                product = models.Product.objects.get(user=user, id=pd_id)
                product.supplier_payment = 0;
                product.save()
                C.products.add(product)
                C.save()
            
        # Customer Set Payment
        if not product_id == None:
            if not supplier_payment == 0:
                S = models.Product.objects.get(id=product_id)
                S.supplier_payment = int(S.supplier_payment) + int(supplier_payment)
                S.save()

        return Response(status=status.HTTP_200_OK)

    def delete(self,request):

        supplier_id = request.GET.get("supplier_id", None)
        products =  request.GET.get('products', None)

        user = get_user_model().objects.get(username=request.user, is_plan=True)

        print(products, supplier_id)

        if not products == None:
            C = models.Supplier.objects.get(id=supplier_id)
            S = models.Product.objects.get(user=user,id=products)
            S.save()
            C.products.remove(S)
            C.save()
        else:
            C = models.Supplier.objects.get(id=supplier_id)
            C.delete()


        return Response(status=status.HTTP_200_OK)



class SoldProduct(APIView):

    def get(self, request):
        rn = request.GET.get['receiptNumber']

        user = get_user_model().objects.get(username=request.user, is_plan=True)
        S = models.Sales.objects.get(user=user, receiptNumber=rn)
        seri = serializers.SoldProductSerializer(S.products.all(), many=True)
        return Response(seri.data)


class TopProductsView(APIView):

    def get(self, request):
        time = request.GET.get('time')
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        d = datetime.now()

        if time == 'today':
            data = models.SoldProduct.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))

        elif time == 'month':
            data = models.SoldProduct.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))

        elif time == 'year':
            data = models.SoldProduct.objects.filter(
                user=user, date__year=str(d.year))

        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            data = models.SoldProduct.objects.filter(
                user=user, date__range=(sd, ed))

        else:
            data = models.SoldProduct.objects.filter(user=user)

        topmoneyproduct = {}
        topfreqsellproduct = {}

        for item in data:
            topmoneyproduct[item.name] = topmoneyproduct.get(
                item.name, 0) + int(float(item.price)) * int(float(item.qty))
            topfreqsellproduct[item.name] = topfreqsellproduct.get(
                item.name, 0) + 1

        CombineData = {
            'T_Money': topmoneyproduct,
            'T_Freq': topfreqsellproduct,
        }

        return Response(CombineData)


def AyearGenerator(self, data, strftime='%b'):
    # Create a list of all months in the year

    if strftime == '%b':
        months = [month_abbr[i] for i in range(1, 13)]
    else:
        months = [month_name[i] for i in range(1, 13)]

    # Initialize the result dictionary with all months set to 0
    result = {month: 0 for month in months}

    # monthString = ['0','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    # 2022-09-06 18:50:44.169216+00:00

    for item in data:
        if len(str(item.date)) == 10:
            # Date string does not include a time component
            d = datetime.strptime(str(item.date), "%Y-%m-%d")
        else:
            # Date string includes a time component
            d = datetime.strptime(str(item.date)[0:19], "%Y-%m-%d %H:%M:%S")

        month_name_str = d.strftime(strftime)
        result[month_name_str] = result.get(
            month_name_str, 0) + int(float(item.price))

    return result


def AmonthGenerator(self, data):
    result = {}
    # monthString = ['0','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    # 2022-09-06 18:50:44.169216+00:00

    for item in data:
        d = datetime.strptime(str(item.date)[0:19], "%Y-%m-%d")
        name = d.strftime('%x')
        result[name] = result.get(name, 0) + int(float(item.price))

    return result


def AtodayGenerator(self, data):
    result = {}
    # monthString = ['0','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    # 2022-09-06 18:50:44.169216+00:00

    for item in data:
        d = datetime.strptime(str(item.date)[0:19], "%Y-%m-%d")
        name = d.strftime('%I:%M %p')
        result[name] = result.get(name, 0) + int(float(item.price))

    return result


def ChartDataGenerator(self, data, time):
    if time == 'today':
        return AtodayGenerator(self, data)
    elif time == 'month':
        return AmonthGenerator(self, data)
    elif time == 'year':
        return AyearGenerator(self, data)
    else:
        return AmonthGenerator(self, data)


class Expense(APIView):
    def get(self, request):
        time = request.GET.get('time')
        d = datetime.now()
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        if time == 'today':
            data = models.Expense.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'month':
            data = models.Expense.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'year':
            data = models.Expense.objects.filter(
                user=user, date__year=str(d.year))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            data = models.Expense.objects.filter(
                user=user, date__range=(sd, ed))
            chartdata = ChartDataGenerator(self, data, time)
        else:
            data = models.Expense.objects.filter(user=user)
            chartdata = ChartDataGenerator(self, data, time)

        s = serializers.ExpenseSerializer(data, many=True)
        CombineData = {
            'DATA': s.data,
            'CHART_LABEL': chartdata.keys(),
            'CHART_DATA': chartdata.values(),
        }

        return Response(CombineData)

    def post(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        title = request.data['title']
        price = request.data['price']
        date = request.data['date']
        description = request.data['description']

        ex = models.Expense.objects.create(
            date=date, user=user, description=description, title=title, price=price)

        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        id = request.data['id']
        title = request.data['title']
        price = request.data['price']
        date = request.data['date']
        description = request.data['description']

        ex = models.Expense.objects.get(user=user, id=id)
        ex.title = title
        ex.price = price
        ex.description = description
        ex.date = date
        ex.save()
        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        id = request.GET.get('id')
        ex = models.Expense.objects.get(user=user, id=id)
        ex.delete()
        return Response(status=status.HTTP_201_CREATED)


class Purchase(APIView):
    def get(self, request):
        time = request.GET.get('time')
        d = datetime.now()
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        data = models.Purchase.objects.filter(user=user)
        if time == 'today':
            data = models.Purchase.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'month':
            data = models.Purchase.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'year':
            data = models.Purchase.objects.filter(
                user=user, date__year=str(d.year))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            data = models.Purchase.objects.filter(
                user=user, date__range=(sd, ed))
            chartdata = ChartDataGenerator(self, data, time)
        else:
            data = models.Purchase.objects.filter(user=user)
            chartdata = ChartDataGenerator(self, data, time)
        s = serializers.PurchaseSerializer(data, many=True)
        CombineData = {
            'DATA': s.data,
            'CHART_LABEL': chartdata.keys(),
            'CHART_DATA': chartdata.values(),
        }

        return Response(CombineData)

    def post(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        title = request.data['title']
        price = request.data['price']
        date = request.data['date']
        description = request.data['description']

        ex = models.Purchase.objects.create(
            date=date, user=user, description=description, title=title, price=price)

        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        id = request.data['id']
        title = request.data['title']
        price = request.data['price']
        date = request.data['date']
        description = request.data['description']

        ex = models.Purchase.objects.get(user=user, id=id)
        ex.title = title
        ex.price = price
        ex.description = description
        ex.date = date
        ex.save()
        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        user = get_user_model().objects.get(username=request.user)
        id = request.GET.get('id')
        ex = models.Purchase.objects.get(user=user, id=id)
        ex.delete()
        return Response(status=status.HTTP_201_CREATED)


class OtherIncome(APIView):
    def get(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        time = request.GET.get('time')

        d = datetime.now()

        if time == 'today':
            data = models.OtherIncome.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))
            chartdata = ChartDataGenerator(self, data, time)
            print(data)
        elif time == 'month':
            data = models.OtherIncome.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'year':
            data = models.OtherIncome.objects.filter(
                user=user, date__year=str(d.year))
            chartdata = ChartDataGenerator(self, data, time)
        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            data = models.OtherIncome.objects.filter(
                user=user, date__range=(sd, ed))
            chartdata = ChartDataGenerator(self, data, time)
        else:
            data = models.OtherIncome.objects.filter(user=user)
            chartdata = ChartDataGenerator(self, data, time)

        s = serializers.OtherIncomeSerializer(data, many=True)
        CombineData = {
            'DATA': s.data,
            'CHART_LABEL': chartdata.keys(),
            'CHART_DATA': chartdata.values(),
        }

        return Response(CombineData)

    def post(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        title = request.data['title']
        price = request.data['price']
        date = request.data['date']
        description = request.data['description']

        ex = models.OtherIncome.objects.create(
            date=date, user=user, description=description, title=title, price=price)

        return Response(status=status.HTTP_201_CREATED)

    def put(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        id = request.data['id']
        title = request.data['title']
        price = request.data['price']
        date = request.data['date']
        description = request.data['description']

        ex = models.OtherIncome.objects.get(user=user, id=id)
        ex.title = title
        ex.price = price
        ex.description = description
        ex.date = date
        ex.save()
        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        id = request.GET.get('id')
        ex = models.OtherIncome.objects.get(user=user, id=id)
        ex.delete()
        return Response(status=status.HTTP_201_CREATED)


class ProfitAndLoss(APIView):
    def get(self, request, format=None):
        user = get_user_model().objects.get(username=request.user, is_plan=True)

        d = datetime.now()

        otherincome_data = models.OtherIncome.objects.filter(
            user=user, date__year=str(d.year))
        sales_data = models.Sales.objects.filter(
            user=user, date__year=str(d.year))
        expense_data = models.Expense.objects.filter(
            user=user, date__year=str(d.year))
        purchase_data = models.Purchase.objects.filter(
            user=user, date__year=str(d.year))

        sales_ge = yearGenerator(self, sales_data, '%B')
        otherincome_ge = AyearGenerator(self, otherincome_data, '%B')
        expense_ge = AyearGenerator(self, expense_data, '%B')
        purchase_ge = AyearGenerator(self, purchase_data, '%B')

        addData = {k: sales_ge.get(k, 0) + otherincome_ge.get(k, 0)
                   for k in set(sales_ge) | set(otherincome_ge)}
        minusData = {k: expense_ge.get(k, 0) + purchase_ge.get(k, 0)
                     for k in set(expense_ge) | set(purchase_ge)}
        subtractData = {k: addData.get(k, 0) - minusData.get(k, 0)
                        for k in set(addData) | set(minusData)}

        print(sales_ge)

        # print(addData,minusData,subtractData)
        ordered_data = sorted(
            addData.items(), key=lambda x: datetime.strptime(x[0], '%B'))
        print(ordered_data)

        CombineData = {
            'addData':  OrderedDict(sorted(addData.items(), key=lambda x: datetime.strptime(x[0], '%B'))),
            'minusData': OrderedDict(sorted(minusData.items(), key=lambda x: datetime.strptime(x[0], '%B'))),
            'result': OrderedDict(sorted(subtractData.items(), key=lambda x: datetime.strptime(x[0], '%B'))),

        }
        return Response(CombineData)


class ProfileAPIView(APIView):

    def get(self, request, format=None):
        user = models.User.objects.get(username=request.user)
        s = serializers.ProfileSerializer(user)

        today = timezone.now()
        # 2022-11-02 08:33:40+00:00
        # endd  = datetime.strptime(str(user.end_d),"%Y-%m-%d %H:%M:%S%z")
        print(today, 'Today')
        endd = user.end_d
        print(endd, 'End Date')
        print(today >= endd, 'Compare Two Date')
        if today >= endd:
            print('end Plan')
            user.is_plan = False
            user.save()
        else:
            user.is_plan = True
            user.save()

        print(timezone.get_current_timezone)

        s = serializers.ProfileSerializer(user)

        return Response(s.data)

    def post(self, request, format=None):
        user = get_user_model().objects.get(username=request.user)

        if 'image' in request.data:
            user.profileimage = compress_image(request.FILES['image'])
            user.save()
            s = serializers.ProfileSerializer(user)
            return Response(s.data)
        s = serializers.ProfileSerializer(user)
        return Response(s.data)

    def delete(self, request, format=None):
        username = request.GET.get('username')
        user = get_user_model().objects.get(username=request.user)

        if user.is_superuser:
            del_user = get_user_model().objects.get(username=username)
            del_user.delete()

            return Response(status=status.HTTP_200_OK)
        return Response('not access')


class FeedBackAPIView(APIView):

    def post(self, request, format=None):
        message = request.data['message']

        models.FeedBack.objects.create(user=request.user, message=message)

        return Response(status=status.HTTP_201_CREATED)


class PricingAPIView(APIView):
    def get(self, request, format=None):
        data = models.Pricing.objects.filter(is_digits=False)
        pricing_ser = serializers.PricingSerializer(data, many=True)
        user = get_user_model().objects.get(username=request.user)
        print(user.is_superuser)
        pr_req_ser = {data: {}}
        try:
            pricing_req = models.PricingRequest.objects.filter(
                user=user, done=False)
            pr_req_ser = serializers.PricingRequestSerializer(
                pricing_req, many=True)
        except ObjectDoesNotExist:
            print('Objects Does Not exist')
        CombineData = {
            'pricing': pricing_ser.data,
            'pr_request': pr_req_ser.data
        }

        return Response(CombineData)

    # It Is Main Buy SOftware not adding Pricing method check this code
    def post(self, request, format=None):
        price_time_type = request.data['type']
        user = get_user_model().objects.get(username=request.user)
        pricing = models.Pricing.objects.get(
            id=price_time_type, is_digits=False)
        models.PricingRequest.objects.create(user=user, rq_price=pricing)

        return Response(status=status.HTTP_201_CREATED)

    def delete(self, request, format=None):
        price_time_type = request.GET.get('type')
        user = get_user_model().objects.get(username=request.user)
        pricing = models.Pricing.objects.get(
            id=price_time_type, is_digits=False)
        pr_req = models.PricingRequest.objects.get(
            user=user, rq_price=pricing, done=False)
        pr_req.delete()

        return Response(status=status.HTTP_201_CREATED)

# Only Super User Can Be Use this View


class PricingRequestView(APIView):

    def get(self, request, format=None):
        user = get_user_model().objects.get(username=request.user)
        if user.is_superuser:
            pricing_req = models.PricingRequest.objects.all()
            ser_p_r = serializers.PricingRequestSerializer(
                pricing_req, many=True)

            return Response(ser_p_r.data)
        return Response('not access')

    def post(self, request):
        handle_user = get_user_model().objects.get(username=request.user)
        if handle_user.is_superuser:
            username = request.data['username']
            rq_id = request.data['rq_id']
            user = get_user_model().objects.get(username=username)
            pr = models.PricingRequest.objects.get(
                id=rq_id, user=user, done=False)
            pr.done = True
            user.is_plan = True
            start_d = datetime.now()
            end = start_d + timedelta(days=int(pr.rq_price.days))
            print(end)
            user.start_d = start_d  # now Date
            user.end_d = end
            user.save()
            pr.save()
            print(user.start_d)

            return Response(status=status.HTTP_201_CREATED)
        return Response('not access')

    def delete(self,request):
        id = request.GET.get('id')
        user = get_user_model().objects.get(username=request.user)
        if user.is_superuser:
            pricing_req = models.PricingRequest.objects.get(id=id)
            pricing_req.delete()

            return Response(status=status.HTTP_200_OK)
        return Response('not access')



class LogoutUserAPIView(APIView):
    queryset = get_user_model().objects.all()

    def get(self, request, format=None):
        device_unique_id = request.GET.get('duid')
        models.Device.objects.filter(unique_id=device_unique_id).delete()
        request.user.auth_token.delete()
        return Response(status=status.HTTP_200_OK)


class LoginWithFacebook(CreateAPIView):
    permission_classes = [AllowAny]
    serializer_class = serializers.CreateUserSerializer

    def post(self, request):

        try:
            username = request.data['username']
            print(username)
            user = models.User.objects.get(username=username)
            print(user)
            token = Token.objects.get(user=user)
            token_data = {'token': token.key}
            return Response({**token_data}, status=status.HTTP_201_CREATED)
        except ObjectDoesNotExist:
            serializers = self.get_serializer(data=request.data)
            serializers.is_valid(raise_exception=True)
            self.perform_create(serializers)

            token = Token.objects.create(user=serializers.instance)
            token_data = {'token': token.key}

            return Response({**token_data}, status=status.HTTP_201_CREATED)


# Write a class that accept the excel file and read excel file add to product model
# Category row is not id in excel file ,it is title so you have to find the category id and if it not have add new category from row before add product


class ExcelUploadNExportAPIView(APIView):  # for products

    def get(self, request, format=None):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        type = request.GET.get('type')
        if type == 'all':
            products = models.Product.objects.filter(user=user)
        elif type == 'available':
            products = models.Product.objects.filter(user=user, qty__gt=0)
        elif type == 'out_of_stock':
            products = models.Product.objects.filter(user=user, qty=0)
        # elif type == 'expired':
        #     products = models.Product.objects.filter(user=user,expiry_date__lt=timezone.now())
        else:
            products = models.Product.objects.filter(user=user)

        # Create a new workbook
        wb = Workbook()

        # Add a worksheet for the data
        ws = wb.active
        ws.title = "Products"

        # Define the columns for the worksheet
        columns = ['Product Name', 'Buy Price', 'Sale Price', 'Qty'
                   'Category', 'Description', 'Barcode']

        # Set the column headers for the worksheet
        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=1, column=col_num,
                    value=column_title).font = Font(bold=True)

        # Add the data to the worksheet
        for row_num, product in enumerate(products, 2):
            ws.cell(row=row_num, column=1, value=product.name)
            ws.cell(row=row_num, column=2, value=product.cost)
            ws.cell(row=row_num, column=2, value=product.price)
            ws.cell(row=row_num, column=3, value=product.qty)
            ws.cell(row=row_num, column=4, value=product.category.title)
            ws.cell(row=row_num, column=5, value=product.description)
            ws.cell(row=row_num, column=6,
                    value=product.barcode)

        # Set the response headers
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Products.xlsx"

        # Save the workbook to the response
        wb.save(response)

        return response

    def post(self, request, format=None):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        file = request.FILES['file']
        print(file)
        # Use openpyxl to read the Excel file
        workbook = load_workbook(file, read_only=True)
        sheet = workbook.active

        # Assuming data starts from the second row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            print(row[0])  # Assuming 'Category' is in the first column
            print(row)
            try:
                category = models.Category.objects.get(
                    title=row[4], user=user)
            except ObjectDoesNotExist:
                category = models.Category.objects.create(
                    title=row[4], user=user)

            models.Product.objects.create(
                name=row[0],  # Assuming 'Name' is in the second column
                cost=row[1],  # Assuming "Buy Price " is in the
                price=row[2],  # Assuming 'Price' is in the third column
                qty=row[3],  # Assuming 'Qty' is in the fourth column
                category=category,
                # Assuming 'Description' is in the fifth column
                description=row[5],
                user=user
            )

        return Response(status=status.HTTP_201_CREATED)


class ExcelExportProfitandLoss(APIView):

    def get(self, request, format=None):
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        today = timezone.now()
        # 2022-11-02 08:33:40+00:00
        # endd  = datetime.strptime(str(user.end_d),"%Y-%m-%d %H:%M:%S%z")

        d = datetime.now()

        print(today, 'Today')
        endd = user.end_d
        print(endd, 'End Date')
        print(today >= endd, 'Compare Two Date')
        if today >= endd:
            print('end Plan')
            user.is_plan = False
            user.save()
        else:
            user.is_plan = True
            user.save()

        print(timezone.get_current_timezone)

        s = serializers.ProfileSerializer(user)

        otherincome_data = models.OtherIncome.objects.filter(
            user=user, date__year=str(d.year))
        sales_data = models.Sales.objects.filter(
            user=user, date__year=str(d.year))
        expense_data = models.Expense.objects.filter(
            user=user, date__year=str(d.year))
        purchase_data = models.Purchase.objects.filter(
            user=user, date__year=str(d.year))

        sales_ge = yearGenerator(self, sales_data, '%B')
        otherincome_ge = AyearGenerator(self, otherincome_data, '%B')
        expense_ge = AyearGenerator(self, expense_data, '%B')
        purchase_ge = AyearGenerator(self, purchase_data, '%B')

        addData = {k: sales_ge.get(k, 0) + otherincome_ge.get(k, 0)
                   for k in set(sales_ge) | set(otherincome_ge)}
        minusData = {k: expense_ge.get(k, 0) + purchase_ge.get(k, 0)
                     for k in set(expense_ge) | set(purchase_ge)}
        subtractData = {k: addData.get(k, 0) - minusData.get(k, 0)
                        for k in set(addData) | set(minusData)}

        print(sales_ge)

        # print(addData,minusData,subtractData)
        ordered_data = sorted(
            addData.items(), key=lambda x: datetime.strptime(x[0], '%B'))
        print(ordered_data)

        CombineData = {
            'addData':  OrderedDict(sorted(addData.items(), key=lambda x: datetime
                                           .strptime(x[0], '%B'))),
            'minusData': OrderedDict(sorted(minusData.items(), key=lambda x: datetime
                                            .strptime(x[0], '%B'))),
            'result': OrderedDict(sorted(subtractData.items(), key=lambda x: datetime
                                         .strptime(x[0], '%B'))),
        }
        # Create a new workbook
        wb = Workbook()

        # Add a worksheet for the data
        ws = wb.active
        ws.title = "ProfitandLoss"

        # Define the columns for the worksheet
        columns = ['Month', 'Income', 'Expense', 'Profit/Loss']

        # Set the column headers for the worksheet
        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=1, column=col_num,
                    value=column_title).font = Font(bold=True)

        # Add the data to the worksheet
        for row_num, month in enumerate(CombineData['addData'], 2):
            ws.cell(row=row_num, column=1, value=month)
            ws.cell(row=row_num, column=2, value=CombineData['addData'][month])
            ws.cell(row=row_num, column=3,
                    value=CombineData['minusData'][month])
            ws.cell(row=row_num, column=4, value=CombineData['result'][month])

        # Set the response headers
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=ProfitandLoss.xlsx"

        # Save the workbook to the response
        wb.save(response)

        return response


# Export All Report includes sales, expenses, purchase, other income to excel file

class ExcelExportAllReportAPIView(APIView):
    def get(self, request, format=None):

        # first get all data report
        type = request.GET.get('type')
        time = request.GET.get('time')
        user = get_user_model().objects.get(username=request.user, is_plan=True)
        d = datetime.now()
        print(d)
        chartdata = {}

        if time == 'today':
            sales_data = models.Sales.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))
            expense_data = models.Expense.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))
            purchase_data = models.Purchase.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))
            otherincome_data = models.OtherIncome.objects.filter(user=user, date__year=str(
                d.year), date__month=str(d.month), date__day=str(d.day))

        elif time == 'month':
            sales_data = models.Sales.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))
            expense_data = models.Expense.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))
            purchase_data = models.Purchase.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))
            otherincome_data = models.OtherIncome.objects.filter(
                user=user, date__year=str(d.year), date__month=str(d.month))
        elif time == 'year':
            sales_data = models.Sales.objects.filter(
                user=user, date__year=str(d.year))
            expense_data = models.Expense.objects.filter(
                user=user, date__year=str(d.year))
            purchase_data = models.Purchase.objects.filter(
                user=user, date__year=str(d.year))
            otherincome_data = models.OtherIncome.objects.filter(
                user=user, date__year=str(d.year))

        elif time == 'custom':
            start_date = request.GET.get('startd')
            end_date = request.GET.get('endd')
            sd = datetime.strptime(start_date, "%m/%d/%y")
            ed = datetime.strptime(
                end_date, "%m/%d/%y").replace(hour=11, minute=59, second=59)
            sales_data = models.Sales.objects.filter(
                user=user, date__range=(sd, ed))
            expense_data = models.Expense.objects.filter(
                user=user, date__range=(sd, ed))
            purchase_data = models.Purchase.objects.filter(
                user=user, date__range=(sd, ed))
            otherincome_data = models.OtherIncome.objects.filter(
                user=user, date__range=(sd, ed))

        else:
            sales_data = models.Sales.objects.filter(user=user)
            expense_data = models.Expense.objects.filter(user=user)
            purchase_data = models.Purchase.objects.filter(user=user)
            otherincome_data = models.OtherIncome.objects.filter(user=user)

        # Make excel file for this all data in different sheet
        # Create a new workbook
        wb = Workbook()

        # Add worksheets for each data type
        sales_ws = wb.create_sheet("Sales")
        expense_ws = wb.create_sheet("Expense")
        otherincome_ws = wb.create_sheet("OtherIncome")

        # Define the columns for each worksheet
        sales_columns = ["Receipt Number", "Customer Name", "Total Amount", "Tax",
                         "Discount", "Delivery Charges", "Grand Total", "Profit", "Date", "Description"]
        expense_columns = ["Title", "Price", "Date", "Description"]
        otherincome_columns = ["Title", "Price", "Date", "Description"]

        # Set the column headers for each worksheet
        for col_num, column_title in enumerate(sales_columns, 1):
            sales_ws.cell(row=1, column=col_num,
                          value=column_title).font = Font(bold=True)
        for col_num, column_title in enumerate(expense_columns, 1):
            expense_ws.cell(row=1, column=col_num,
                            value=column_title).font = Font(bold=True)

        for col_num, column_title in enumerate(otherincome_columns, 1):
            otherincome_ws.cell(row=1, column=col_num,
                                value=column_title).font = Font(bold=True)

        # Add the data to each worksheet
        for row_num, sale in enumerate(sales_data, 2):
            sales_ws.cell(row=row_num, column=1, value=sale.voucherNumber)
            sales_ws.cell(row=row_num, column=2, value=sale.customerName)
            sales_ws.cell(row=row_num, column=3, value=sale.totalAmount)
            sales_ws.cell(row=row_num, column=4, value=sale.tax)
            sales_ws.cell(row=row_num, column=5, value=sale.discount)
            sales_ws.cell(row=row_num, column=6, value=sale.deliveryCharges)
            sales_ws.cell(row=row_num, column=7, value=sale.grandtotal)
            sales_ws.cell(row=row_num, column=8, value=sale.totalProfit)
            sales_ws.cell(row=row_num, column=9,
                          value=sale.date.strftime("%m-%d-%Y"))
            sales_ws.cell(row=row_num, column=10, value=sale.description)

        for row_num, expense in enumerate(expense_data, 2):
            expense_ws.cell(row=row_num, column=1, value=expense.title)
            expense_ws.cell(row=row_num, column=2, value=expense.price)
            expense_ws.cell(row=row_num, column=3,
                            value=expense.date.strftime("%m-%d-%Y"))
            expense_ws.cell(row=row_num, column=4, value=expense.description)

        for row_num, otherincome in enumerate(otherincome_data, 2):
            otherincome_ws.cell(row=row_num, column=1, value=otherincome.title)
            otherincome_ws.cell(row=row_num, column=2, value=otherincome.price)
            otherincome_ws.cell(row=row_num, column=3,
                                value=otherincome.date.strftime("%m-%d-%Y"))
            otherincome_ws.cell(row=row_num, column=4,
                                value=otherincome.description)

        # Set the response headers
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=AllReport.xlsx"

        # Save the workbook to the response
        wb.save(response)

        return response


def generate_barcode(product_id):

    product_id = str(product_id).zfill(12)
    # print(product_id, 'what.....')
    # Generate a unique barcode value using the EAN13 format
    ean = barcode.get_barcode_class('ean13')
    barcode_value = ean(str(product_id), writer=ImageWriter())

    # Convert the barcode image to a base64-encoded string
    buffer = io.BytesIO()
    barcode_value.write(buffer)
    barcode_image = Image.open(buffer)
    return barcode_image

# Export BarCode for all products


class ExcelExportBarCodeAPIView(APIView):
    def get(self, request, format=None):
        user = get_user_model().objects.get(username=request.user, is_plan=True)

        pdlist = request.GET.get('sid')
        pdlist = ast.literal_eval(pdlist)
        product_ids = [int(id) for id in pdlist]
        products = models.Product.objects.filter(user=user, id__in=product_ids)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Products_BarCode.pdf"'

        # Create a buffer to hold the PDF file
        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)

        # Set the column and row dimensions
        col_width = 40 * mm
        row_height = 20 * mm
        margin = 10 * mm
        x = margin
        y = A4[1] - margin

        current_product_name = None

        for i in products:
            for j in range(int(i.qty)):
                barcode_image = generate_barcode(i.barcode)

                # Resize the barcode image to fit in the cell
                # barcode_image = barcode_image.resize((int(col_width), int(row_height)))

                # Convert the barcode image to a format that can be added to the PDF
                img_data = io.BytesIO()
                barcode_image.save(img_data, format='PNG')
                img_data.seek(0)

                # c.rect(x, y - row_height - 15 * mm, col_width, row_height, stroke=1, fill=0)

                # Add the barcode image to the PDF
                c.drawImage(ImageReader(img_data), x, y - row_height,
                            width=col_width, height=row_height)

                # Move to the next cell
                x += col_width + 10 * mm

                # If we reach the end of the row, move to the next row
                if x >= A4[0] - margin:
                    x = margin
                    y -= row_height + 10 * mm

                # If we reach the end of the page, start a new page
                if y <= margin:
                    c.showPage()
                    y = A4[1] - margin
                    current_product_name = None

        c.save()

        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response
